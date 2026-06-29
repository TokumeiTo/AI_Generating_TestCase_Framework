import base64
import io
import openpyxl

class ExcelParserService:
    def __init__(self):
        pass

    async def parse_and_filter_spec(self, base64_data: str, target_keyword: str) -> str:
        """
        Decodes base64 excel files and extracts ONLY the header row and rows matching 
        the target keyword to safely minimize token usage. Fallbacks cleanly if keyword not found.
        """
        # 1. Decode the Base64 string back into a byte stream
        file_bytes = base64.b64decode(base64_data)
        excel_stream = io.BytesIO(file_bytes)
        
        # 2. Open workbook in-memory
        wb = openpyxl.load_workbook(excel_stream, data_only=True)
        compiled_lines = []
        keyword = target_keyword.strip()
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_rows = []
            keyword_found = False
            
            # Read rows into memory as strings
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    row_cells = [str(cell).strip() if cell is not None else "" for cell in row]
                    row_text = "\t".join(row_cells).strip()
                    sheet_rows.append(row_text)
                    
                    if keyword in row_text:
                        keyword_found = True
            
            # Extract just the targeted keyword blocks from this sheet
            if keyword_found:
                compiled_lines.append(f"--- Sheet: {sheet_name} ---")
                if sheet_rows:
                    compiled_lines.append(sheet_rows[0]) # Header Row
                
                for r_text in sheet_rows[1:]:
                    if keyword in r_text:
                        compiled_lines.append(r_text)
                        
        # 3. Micro-Fallback implementation if keyword was completely absent
        if not compiled_lines:
            for sheet_name in wb.sheetnames[:1]:
                sheet = wb[sheet_name]
                compiled_lines.append(f"--- Sheet Sample (Keyword Not Found): {sheet_name} ---")
                for row in list(sheet.iter_rows(values_only=True))[:15]:
                    if any(row):
                        row_cells = [str(cell).strip() if cell is not None else "" for cell in row]
                        compiled_lines.append("\t".join(row_cells).strip())
                        
        return "\n".join(compiled_lines)